from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.http import HttpResponse
from .models import Prenda, Gasto
from .forms import GastoForm, ImportarPrendasForm, PrendaForm
import openpyxl
from decimal import Decimal
from django.db import transaction
from django.db.models import Q, Sum


@login_required
def lista_prendas(request):
    estado = request.GET.get('estado', 'todas')

    busqueda = request.GET.get('busqueda', '')
    filtro_talla = request.GET.get('talla', '')
    filtro_marca = request.GET.get('marca', '')
    orden = request.GET.get('orden', '')
    vista = request.GET.get('vista', 'tarjetas')
    if vista not in {'tarjetas', 'tabla'}:
        vista = 'tarjetas'

    prendas = Prenda.objects.all()
    if estado != 'todas':
        prendas = prendas.filter(estado=estado)
    if busqueda:
        prendas = prendas.filter(
            Q(tipo_de_prenda__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(color__icontains=busqueda) |
            Q(localizador__icontains=busqueda)
        )
    if filtro_talla:
        prendas = prendas.filter(talla=filtro_talla)
    if filtro_marca:
        prendas = prendas.filter(marca__icontains=filtro_marca)
    if orden == 'precio_asc':
        prendas = prendas.order_by('precio_vendido')
    elif orden == 'precio_desc':
        prendas = prendas.order_by('-precio_vendido')
    elif orden == 'antiguas':
        prendas = prendas.order_by('id')

    tallas = Prenda.objects.values_list('talla', flat=True).distinct()
    marcas = Prenda.objects.values_list('marca', flat=True).distinct()

    total = Prenda.objects.count()
    vendidas = Prenda.objects.filter(estado='vendido').count()
    disponibles = Prenda.objects.filter(estado='disponible').count()
    borradores = Prenda.objects.filter(estado='borrador').count()

    paginator = Paginator(prendas, 8)
    page_number = request.GET.get('page')
    prendas = paginator.get_page(page_number)

    form = PrendaForm()
    if request.method == 'POST':
        form = PrendaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Prenda añadida correctamente.')
            return redirect('lista_prendas')

    prendas_vendidas = Prenda.objects.filter(estado='vendido')
    beneficio_total = sum(
        (p.beneficio() for p in prendas_vendidas if p.beneficio() is not None), Decimal('0.00'))
    facturacion_total = prendas_vendidas.aggregate(total=Sum('precio_vendido'))['total'] or Decimal('0.00')
    return render(request, 'prendas/lista_prendas.html', {
        'prendas': prendas,
        'estado': estado,
        'total': total,
        'vendidas': vendidas,
        'disponibles': disponibles,
        'borradores': borradores,
        'form': form,
        'beneficio_total': beneficio_total,
        'facturacion_total': facturacion_total,
        # El resultado neto del inventario compara solo ventas y coste de
        # compra de las prendas que ya se han vendido.
        'resultado_neto': beneficio_total,
        'busqueda': busqueda,
        'filtro_talla': filtro_talla,
        'filtro_marca': filtro_marca,
        'orden': orden,
        'vista': vista,
        'tallas': tallas,
        'marcas': marcas,
    })


@login_required
def editar_prenda(request, pk):
    prenda = get_object_or_404(Prenda, pk=pk)
    form = PrendaForm(instance=prenda)

    if request.method == 'POST':
        form = PrendaForm(request.POST, instance=prenda)
        if form.is_valid():
            form.save()
            messages.success(request, 'Prenda actualizada correctamente.')
            return redirect('lista_prendas')

    return render(request, 'prendas/editar_prenda.html', {'form': form, 'prenda': prenda})


@login_required
def eliminar_prenda(request, pk):
    prenda = get_object_or_404(Prenda, pk=pk)
    if request.method == 'POST':
        prenda.delete()
        messages.success(request, 'Prenda eliminada.')
        return redirect('lista_prendas')
    return render(request, 'prendas/eliminar_prenda.html', {'prenda': prenda})


@login_required
def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prendas"

    ws.append(['Tipo', 'Talla', 'Color', 'Marca', 'Localizador', 'Donde subido', 'Precio comprado', 'Precio vendido', 'Beneficio', 'Estado', 'Fecha de alta'])

    for prenda in Prenda.objects.all():
        ws.append([
            prenda.tipo_de_prenda,
            prenda.talla,
            prenda.color,
            prenda.marca,
            prenda.localizador,
            prenda.donde_esta_subido,
            float(prenda.precio_comprado),
            float(prenda.precio_vendido) if prenda.precio_vendido else '',
            float(prenda.beneficio()) if prenda.beneficio() is not None else '',
            prenda.get_estado_display(),
            prenda.fecha_creacion.strftime('%d/%m/%Y'),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="prendas.xlsx"'
    wb.save(response)
    return response


@login_required
def importar_prendas(request):
    form = ImportarPrendasForm()
    if request.method == 'POST':
        form = ImportarPrendasForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                libro = openpyxl.load_workbook(form.cleaned_data['archivo'], read_only=True, data_only=True)
                hoja = libro.active
                filas = hoja.iter_rows(values_only=True)
                cabeceras = next(filas, None)
                if not cabeceras:
                    raise ValueError('El Excel está vacío.')

                indices = {str(nombre).strip().lower(): indice for indice, nombre in enumerate(cabeceras) if nombre}
                requeridas = {'tipo', 'talla', 'color', 'marca', 'donde subido', 'precio comprado', 'estado'}
                faltantes = requeridas - indices.keys()
                if faltantes:
                    raise ValueError('No parece ser un archivo exportado por la aplicación.')

                estados = {'borrador': 'borrador', 'disponible': 'disponible', 'vendido': 'vendido'}
                prendas = []
                for numero, fila in enumerate(filas, start=2):
                    if not any(valor is not None and str(valor).strip() for valor in fila):
                        continue
                    try:
                        estado = estados.get(str(fila[indices['estado']]).strip().lower())
                        if not estado:
                            raise ValueError('estado no válido')
                        precio_comprado = Decimal(str(fila[indices['precio comprado']]))
                        precio_vendido = None
                        indice_venta = indices.get('precio vendido')
                        if indice_venta is not None and fila[indice_venta] not in (None, ''):
                            precio_vendido = Decimal(str(fila[indice_venta]))
                        prendas.append(Prenda(
                            tipo_de_prenda=str(fila[indices['tipo']]).strip(),
                            talla=str(fila[indices['talla']]).strip(),
                            color=str(fila[indices['color']]).strip(),
                            marca=str(fila[indices['marca']]).strip(),
                            localizador=str(fila[indices.get('localizador', -1)] or '').strip() if 'localizador' in indices else '',
                            donde_esta_subido=str(fila[indices['donde subido']]).strip(),
                            precio_comprado=precio_comprado,
                            precio_vendido=precio_vendido,
                            estado=estado,
                        ))
                    except (ArithmeticError, IndexError, TypeError, ValueError) as error:
                        raise ValueError(f'La fila {numero} no es válida ({error}).') from error

                if not prendas:
                    raise ValueError('No se han encontrado prendas para importar.')
                with transaction.atomic():
                    Prenda.objects.bulk_create(prendas)
                messages.success(request, f'Se han importado {len(prendas)} prendas correctamente.')
                return redirect('lista_prendas')
            except (ValueError, openpyxl.utils.exceptions.InvalidFileException) as error:
                form.add_error('archivo', str(error))

    return render(request, 'prendas/importar_prendas.html', {'form': form})


@login_required
def lista_gastos(request):
    form = GastoForm()
    if request.method == 'POST':
        form = GastoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gasto añadido correctamente.')
            return redirect('lista_gastos')

    gastos = Gasto.objects.all()
    total_gastos = sum(g.importe for g in gastos)

    context = {
        'gastos': gastos,
        'total_gastos': total_gastos,
        'form': form,
    }
    return render(request, 'prendas/gastos.html', context)


@login_required
def eliminar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    if request.method == 'POST':
        gasto.delete()
        messages.success(request, 'Gasto eliminado.')
        return redirect('lista_gastos')
    return render(request, 'prendas/eliminar_gasto.html', {'gasto': gasto})
